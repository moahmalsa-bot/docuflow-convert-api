"""
Excel -> PDF using LibreOffice Calc headless (the real spreadsheet print
engine). Existing print settings are respected exactly; sane defaults are
applied ONLY to sheets that have none.

Process lifecycle: every conversion runs with its OWN unique LibreOffice
profile (never the shared default — a locked profile is what makes soffice
hang forever), under a strict 60-second timeout that kills the whole process
tree, serialized to ONE conversion at a time. Stale soffice processes older
than 2 minutes are terminated before starting.

Output lifecycle: the safe filename is computed EXACTLY ONCE
(sanitize(original_stem) + ".pdf", trailing dots stripped so "Report..xlsx"
becomes "Report.pdf" — never a double extension). The PDF is copied to
/app/storage/outputs/{job_id}/{safe_file_name} and verified BEFORE success is
returned. The download endpoint resolves ONLY that path on disk — no
in-memory registry — and never deletes the file, so repeated downloads work.
24h retention.
"""

import hashlib
import logging
import os
import re
import shutil
import signal
import subprocess
import threading
import time
import uuid
import zipfile
from pathlib import Path
from urllib.parse import quote

from fastapi import HTTPException, Request
from pypdf import PdfReader
from starlette.responses import FileResponse, JSONResponse

from app.utils.files import cleanup_old_directories, libreoffice_binary, sanitize_filename

logger = logging.getLogger("docuflow.excel_pdf")

# Permanent runtime storage — NEVER a TemporaryDirectory. Created at startup.
OUTPUT_ROOT = Path(os.getenv("OUTPUT_DIR", "/app/storage/outputs"))
OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

# Completed files are kept for 24 hours; cleanup deletes ONLY older files.
OUTPUT_MAX_AGE_SECONDS = int(os.getenv("OUTPUT_MAX_AGE_SECONDS", "86400"))
# Strict per-process timeout — soffice is killed (whole tree) beyond this.
EXCEL_TIMEOUT_SECONDS = int(os.getenv("EXCEL_TIMEOUT_SECONDS", "60"))
STALE_SOFFICE_SECONDS = 120

DOWNLOAD_MIME_TYPES = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".zip": "application/zip",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
}

_active_digests: set[str] = set()
_active_lock = threading.Lock()
# ONE LibreOffice conversion at a time on the small Render instance.
_soffice_lock = threading.Lock()


class ExcelTimeout(Exception):
    """LibreOffice exceeded the strict timeout and was killed."""


def _safe_pdf_name(stem: str) -> str:
    """The ONE place a download filename is built: sanitize the stem, strip
    trailing dots/spaces (so "Report." never becomes "Report..pdf"), append
    ".pdf". Idempotent — re-deriving from the returned name yields the same
    name, so the download endpoint resolves the identical stored path."""
    name = sanitize_filename(stem).strip().rstrip(". ").strip() or "converted"
    return f"{name}.pdf"


def _kill_stale_soffice(max_age_seconds: int = STALE_SOFFICE_SECONDS) -> None:
    """Terminates ONLY soffice processes older than 2 minutes. A currently
    active conversion (younger than that) is never touched."""
    now = time.time()
    for proc_dir in Path("/proc").glob("[0-9]*"):
        try:
            cmdline = (proc_dir / "cmdline").read_bytes().replace(b"\x00", b" ").decode(errors="ignore")
            if "soffice" not in cmdline:
                continue
            age = now - proc_dir.stat().st_mtime
            if age > max_age_seconds:
                pid = int(proc_dir.name)
                logger.warning("killing stale soffice: pid=%s age=%.0fs", pid, age)
                os.kill(pid, signal.SIGKILL)
        except (OSError, ValueError):
            continue


def _run_soffice(cmd: list[str], profile_dir: Path) -> None:
    """Runs soffice in its own session with full logging. On timeout the
    ENTIRE process tree is killed and the temp profile is deleted."""
    start = time.time()
    logger.info("soffice start: time=%s cmd=%s", time.strftime("%Y-%m-%dT%H:%M:%S"), " ".join(cmd))
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, start_new_session=True)
    logger.info("soffice pid=%s", proc.pid)
    try:
        stdout, stderr = proc.communicate(timeout=EXCEL_TIMEOUT_SECONDS)
    except subprocess.TimeoutExpired:
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except OSError:
            proc.kill()
        proc.wait()
        shutil.rmtree(profile_dir, ignore_errors=True)
        logger.error("soffice TIMEOUT killed: pid=%s elapsed=%.1fs", proc.pid, time.time() - start)
        raise ExcelTimeout()
    elapsed = time.time() - start
    out = stdout.decode(errors="ignore").strip()
    err = stderr.decode(errors="ignore").strip()
    logger.info(
        "soffice done: pid=%s rc=%s elapsed=%.1fs stdout=%s stderr=%s",
        proc.pid, proc.returncode, elapsed, out[:500], err[:500],
    )
    if proc.returncode != 0:
        raise HTTPException(status_code=500, detail=f"LibreOffice failed (rc={proc.returncode}): {err[:400]}")


def _workbook_has_drawings(path: Path) -> bool:
    """Charts/images/drawings present -> openpyxl round-trip would drop them,
    so the workbook must never be rewritten."""
    try:
        with zipfile.ZipFile(path) as z:
            return any(n.startswith(("xl/charts/", "xl/media/", "xl/drawings/")) for n in z.namelist())
    except zipfile.BadZipFile:
        return True


def _sheet_is_wide(ws) -> bool:
    total_width = 0.0
    for col in range(1, ws.max_column + 1):
        letter = ws.cell(row=1, column=col).column_letter
        dim = ws.column_dimensions.get(letter)
        total_width += dim.width if dim and dim.width else 8.43
    total_height = sum(
        (ws.row_dimensions[r].height if r in ws.row_dimensions and ws.row_dimensions[r].height else 15.0)
        for r in range(1, min(ws.max_row, 200) + 1)
    )
    return total_width * 7 > max(total_height * (4 / 3), 700)


def _apply_default_print_settings(path: Path) -> None:
    """Sheets that already have a print area / fit-to-page / scale are left
    untouched. Only unconfigured sheets get: used-range print area, A4,
    landscape when wide, fit one page wide (unlimited tall), centred."""
    if path.suffix.lower() not in {".xlsx", ".xlsm"} or _workbook_has_drawings(path):
        return
    import openpyxl

    wb = openpyxl.load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    changed = False
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        fit_cfg = bool(ws.sheet_properties.pageSetUpPr and ws.sheet_properties.pageSetUpPr.fitToPage)
        has_config = bool(ws.print_area) or fit_cfg or bool(ws.page_setup.scale and ws.page_setup.scale != 100)
        if has_config or ws.max_row < 1 or ws.max_column < 1:
            continue
        ws.print_area = ws.calculate_dimension()  # real used range only
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "landscape" if _sheet_is_wide(ws) else "portrait"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # allow multiple pages vertically
        ws.print_options.horizontalCentered = True
        changed = True
    if changed:
        wb.save(path)


def _count_visible_sheets(path: Path) -> int:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True)
        count = sum(1 for ws in wb.worksheets if ws.sheet_state == "visible")
        wb.close()
        return count
    import xlrd

    book = xlrd.open_workbook(str(path), on_demand=True)
    return sum(1 for i in range(book.nsheets) if book.sheet_visible(i) == 0) or book.nsheets


def _page_is_blank(page) -> bool:
    if (page.extract_text() or "").strip():
        return False
    resources = page.get("/Resources") or {}
    return "/XObject" not in resources  # no text and no drawn objects


def excel_to_pdf_job(input_file: Path, job_dir: Path, request: Request) -> dict:
    digest = hashlib.sha256(input_file.read_bytes()).hexdigest()
    with _active_lock:
        if digest in _active_digests:
            raise HTTPException(status_code=429, detail="This file is already being converted.")
        _active_digests.add(digest)
    profile_dir = Path(f"/tmp/lo-profile-{uuid.uuid4().hex}")
    try:
        sheet_count = _count_visible_sheets(input_file)
        if sheet_count == 0:
            raise HTTPException(status_code=400, detail="The workbook has no visible worksheets.")

        _apply_default_print_settings(input_file)

        # Unique per-job profile — the shared default profile is NEVER reused
        # (a leftover lock on it makes soffice hang forever).
        cmd = [
            libreoffice_binary(),
            "--headless",
            "--nologo",
            "--nodefault",
            "--nofirststartwizard",
            "--nolockcheck",
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to",
            "pdf:calc_pdf_Export",
            "--outdir",
            str(job_dir),
            str(input_file),
        ]
        _kill_stale_soffice()
        try:
            with _soffice_lock:  # one conversion at a time
                _run_soffice(cmd, profile_dir)
        except ExcelTimeout:
            return JSONResponse(
                status_code=504,
                content={"success": False, "error": "Excel conversion timed out after 60 seconds."},
            )

        pdf_path = job_dir / f"{input_file.stem}.pdf"
        if not pdf_path.exists():
            matches = list(job_dir.glob("*.pdf"))
            if not matches:
                raise HTTPException(status_code=500, detail="LibreOffice did not produce a PDF.")
            pdf_path = matches[0]

        # ---- output validation ----
        if pdf_path.stat().st_size <= 0:
            raise HTTPException(status_code=500, detail="Conversion produced an empty PDF.")
        reader = PdfReader(str(pdf_path))
        page_count = len(reader.pages)
        if page_count == 0:
            raise HTTPException(status_code=500, detail="Conversion produced a PDF with no pages.")
        if page_count < sheet_count:
            raise HTTPException(
                status_code=500,
                detail=f"Only {page_count} page(s) exported for {sheet_count} visible worksheet(s).",
            )
        if _page_is_blank(reader.pages[0]) or _page_is_blank(reader.pages[-1]):
            raise HTTPException(status_code=500, detail="The exported PDF has a blank first or last page.")

        # ---- COPY into permanent storage BEFORE the temp dir is deleted ----
        # The safe filename is built EXACTLY ONCE here and never renamed or
        # re-sanitized differently again. Cleanup removes ONLY jobs > 24h old.
        OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
        cleanup_old_directories(OUTPUT_ROOT, OUTPUT_MAX_AGE_SECONDS)
        job_id = uuid.uuid4().hex
        safe_file_name = _safe_pdf_name(input_file.stem)
        final_dir = OUTPUT_ROOT / job_id
        final_dir.mkdir(parents=True, exist_ok=True)
        final_path = (final_dir / safe_file_name).resolve()
        shutil.copy2(pdf_path, final_path)

        # success is returned ONLY after the permanent copy is verified
        exists = final_path.exists() and final_path.is_file()
        size = final_path.stat().st_size if exists else 0
        logger.info(
            "CONVERSION OUTPUT: job_id=%s generated_path=%s final_path=%s file_name=%s exists=%s size=%d",
            job_id, pdf_path, final_path, safe_file_name, exists, size,
        )
        if not exists or size <= 0:
            raise HTTPException(status_code=500, detail="Converted PDF could not be stored for download.")

        base = str(request.base_url).rstrip("/")
        if base.startswith("http://") and request.headers.get("x-forwarded-proto") == "https":
            base = "https://" + base[len("http://"):]
        return {
            "success": True,
            "fileName": safe_file_name,
            "downloadUrl": f"{base}/download/{job_id}/{quote(safe_file_name)}",
            "sheetCount": sheet_count,
            "pageCount": page_count,
            "fileSize": size,
        }
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)
        with _active_lock:
            _active_digests.discard(digest)


def _not_found() -> JSONResponse:
    return JSONResponse(
        status_code=404,
        content={"success": False, "error": "Converted file was not found or has expired."},
    )


def excel_download_response(job_id: str, file_name: str):
    """Resolves ONLY /app/storage/outputs/{job_id}/{safe_file_name} on disk —
    no in-memory registry, no temp folders. Legitimate names with dots (e.g.
    "19.07.2026 Report.pdf") are accepted; traversal is prevented by the
    resolved-path containment check, NOT by rejecting dots. The file is never
    deleted after download, so pressing Download twice works."""
    if not re.fullmatch(r"[0-9a-f]{32}", job_id) or "/" in file_name or "\\" in file_name:
        logger.warning("download rejected: job_id=%s file_name=%s", job_id, file_name)
        return _not_found()

    # Same single sanitization as at save time -> identical stored path.
    safe_name = _safe_pdf_name(Path(file_name).stem) if file_name.lower().endswith(".pdf") else sanitize_filename(file_name)
    path = (OUTPUT_ROOT / job_id / safe_name).resolve()
    if not str(path).startswith(str(OUTPUT_ROOT.resolve()) + os.sep):
        logger.warning("download rejected (outside storage): job_id=%s path=%s", job_id, path)
        return _not_found()

    exists = path.is_file()
    size = path.stat().st_size if exists else 0
    logger.info(
        "DOWNLOAD REQUEST: job_id=%s requested_file_name=%s resolved_path=%s exists=%s size=%d",
        job_id, file_name, path, exists, size,
    )
    if not exists or size <= 0:
        return _not_found()

    media_type = DOWNLOAD_MIME_TYPES.get(path.suffix.lower(), "application/octet-stream")
    ascii_name = path.name.encode("ascii", "replace").decode()
    disposition = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(path.name)}"
    return FileResponse(
        path=path,
        media_type=media_type,
        filename=path.name,
        headers={
            "Content-Disposition": disposition,
            "Content-Length": str(size),
            "Cache-Control": "no-store",
        },
    )
